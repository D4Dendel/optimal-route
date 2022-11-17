import tkinter as tk
import pyexcel
import tkinter.ttk as ttk
import collections
from tkinter import*
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from collections import defaultdict



root=tk.Tk()
root.title('KOLEKTA: Garbage Collection System')
#root.iconbitmap('D:\THESIS\CODE\klkta.ico')


#root.minsize(height=500, width=800)
canvas = tk.Canvas(root,  width=600, height=300)
canvas.grid(columnspan=3, rowspan=3)

#start_point = 0
# This class represents a directed graph using
# adjacency list representation
dfs_paths = []
class Graph:
 
    # Constructor
    def __init__(self):
 
        # default dictionary to store graph
        self.graph = defaultdict(list)
 
    # function to add an edge to graph
    def addEdge(self, u, v):
        self.graph[u].append(v)
 
    # A function used by DFS
    def DFSUtil(self, v, visited):
 
        # Mark the current node as visited
        # and print it
        visited.add(v)
        print(v, end=' ')
        #print("Visited", v)
        dfs_paths.append(v)
 
        # Recur for all the vertices
        # adjacent to this vertex
        for neighbour in self.graph[v]:
            if neighbour not in visited:
                self.DFSUtil(neighbour, visited)
 
    # The function to do DFS traversal. It uses
    # recursive DFSUtil()
    def DFS(self, v):
 
        # Create a set to store visited vertices
        visited = set()
 
        # Call the recursive helper function
        # to print DFS traversal
        self.DFSUtil(v, visited)


def tab1():
	def tab2():
		#OPENFILE
		file = "D:\COMSCI\Kolekta\Kolekta Revision 1\FINAL-DATA-KOLEKTA.xlsx"
		browse_text = tk.StringVar()

		frame.destroy()
		instructions.destroy()
		button1.destroy()
		button_dfs.destroy()
		scrollbar.destroy()
		#combodata.destroy()
		label2=Label(root, text='OPTIMAL ROUTE FOR BFS', font=('Calibri',25, 'bold'))
		label2.grid(column=1, row=1, sticky = N)
		# BFS algorithm
		paths = []
		def bfs(graph, root):

		    visited, queue = set(), collections.deque([root])
		    visited.add(root)

		    while queue:

		        # Dequeue a vertex from queue
		        vertex = queue.popleft()
		        print(str(vertex) + " ", end="")
		        paths.append(vertex)
		        #print(elements[vertex])

		        # If not visited, mark it as visited, and
		        # enqueue it
		        for neighbour in graph[vertex]:
		            if neighbour not in visited:
		                visited.add(neighbour)
		                queue.append(neighbour)
		if __name__ == '__main__':
			print("Following is Breadth First Traversal: ")
			graph={	0:[1],
					1:[2],
					2:[3],
					3:[4],
					4:[5],
					5:[10,6],
					6:[7,11],
					7:[8],
					8:[9,10],
					9:[11,7],
					10:[6],
					11:[12],
					12:[9]
					}	


			#print ("bfsstart_point", 0)
			bfs(graph, 0)
			#text box
			list = ''
			print ("Test paths", paths)
			for xxx in paths:
				print ("xxx", xxx)
				#path = paths[xxx]
				dash = '	'

				print(elements[xxx])
				list = f'{list+str(xxx)+dash+str(elements[xxx])}\n'


			text_box = tk.Text(root, height=10, width=30, padx=10, pady=10, font='Calibri', fg='black',bg='white')
			text_box.insert(1.0, list)
			text_box.tag_configure("center", justify="left")
			text_box.tag_add("center", 1000.0, "end")
			text_box.grid(column=1, row=2)
			
			# Create a scrollbar
			scroll_bar = tk.Scrollbar(root)
			text_box.configure(yscrollcommand=scroll_bar.set)
			scroll_bar.grid(row = 2, column = 1, sticky = "nse")

			browse_text.set("Browse")

		def back():
			label2.destroy()
			button2.destroy()
			#combodata.destroy()
			text_box.destroy()
			scroll_bar.destroy()
			tab1()
		button2=Button(root, text='BACK',font=('Calibri',20),command=back, bg="#be204a", fg="white")
		button2.grid(column=1,row=3)
		## END OF TAB 222

	def tab3():
		#OPENFILE
		file = "D:\COMSCI\Kolekta\Kolekta Revision 1\FINAL-DATA-KOLEKTA.xlsx"
		browse_text = tk.StringVar()

		#logo_label.destroy()
		frame.destroy()
		instructions.destroy()
		button1.destroy()
		button_dfs.destroy()
		#combodata.destroy()
		label3=Label(root, text='OPTIMAL ROUTE FOR DFS', font=('Calibri',25, 'bold'))
		label3.grid(column=1, row=1, sticky = N)
		
		# Driver code
		# Create a graph given
		# in the above diagram
		g = Graph()
		g.addEdge(0,1)
		g.addEdge(1,2)
		g.addEdge(2,3)
		g.addEdge(3,4)
		g.addEdge(4,5)
		g.addEdge(5,10)
		g.addEdge(5,6)
		g.addEdge(6,7)
		g.addEdge(6,11)
		g.addEdge(7,8)
		g.addEdge(8,9)
		g.addEdge(8,10)
		g.addEdge(9,11)
		g.addEdge(9,7)
		g.addEdge(10,6)
		g.addEdge(11,12)
		g.addEdge(12,9)


		print("Following is DFS from (starting from vertex 0)")
		#print ("dfsstart_point", start_point)
		g.DFS(0)
		#text box
		
		#print ("Test paths", dfs_)
		list = ''
		for xxx in dfs_paths:
			print ("xxx", xxx)
			#path = paths[xxx]
			dash = '	'
			print(elements[xxx])
			list = f'{list+str(xxx)+dash+str(elements[xxx])}\n'
		text_box = tk.Text(root, height=10, width=30, padx=10, pady=10, font='Calibri', fg='black',bg='white')
		text_box.insert(1.0, list)
		text_box.tag_configure("center", justify="left")
		text_box.tag_add("center", 1000.0, "end")
		text_box.grid(column=1, row=2)

		# Create a scrollbar
		scroll_bar = tk.Scrollbar(root)
		text_box.configure(yscrollcommand=scroll_bar.set)
		scroll_bar.grid(row = 2, column = 1, sticky = "nse")

		

		browse_text.set("Browse")

		def back():
			label3.destroy()
			button4.destroy()
			#combodata.destroy()
			text_box.destroy()
			scroll_bar.destroy()
			tab1()
		button4=Button(root, text='BACK',font=('Times_New_Roman',20),command=back, bg="#be204a", fg="white")
		button4.grid(column=1,row=3)
	
		## END OF TAB 3
	#label1=Label(root, text='THIS IS FIRST TAB', font=('Times_New_Roman',25))
	#label1.grid(column=1,row=0)

	#logo

	logo = Image.open('D:\COMSCI\Kolekta\Kolekta Revision 1\logo.png') #resize ni yots hahaha 200x200
	resize_image = logo.resize((450, 300)) 
	logo = ImageTk.PhotoImage(resize_image) 
	logo_label = tk.Label(image=logo)
	logo_label.image = logo
	logo_label.grid(column=1,row=0)

	#instructions
	instructions = tk.Label(root, text="PICK UP AREAS", font=("Raleway", 20, 'bold'))
	instructions.grid(columnspan=3, column=0, row=1, sticky=S)

	# DROP POINTS LIST
	file = "D:\COMSCI\Kolekta\Kolekta Revision 1\FINAL-DATA-KOLEKTA.xlsx"
	read_excel = load_workbook(file)
	ws = read_excel['DATA']
	xlsx_range = ws['I3':'I24']
	elements = []
	for cell in xlsx_range:
		for x in cell:
			y = x.value
			elements.append(y)
			#print(y)

	'''clicked = StringVar()
	combodata = ttk.Combobox(root,state='readonly', values=elements,font="Raleway")
	combodata.current(0)
	combodata.grid(column=1,row=1,sticky=S)'''
	#--------------------------------------

	#DROP POINTS INFORMATION
	

	frame = tk.Frame(root, pady = 10, padx = 10)
	frame.grid(column=1,row=2)
	text = tk.Text(frame, width = 40, height =  10, font = ("Calibri", 10))
	text.grid(column=1,row=2, rowspan = 6, columnspan = 2, pady = 5, padx = 5)
	scrollbar = tk.Scrollbar(frame, command = text.yview)
	text['yscroll'] = scrollbar.set
	scrollbar.grid(column=6,row=2, rowspan = 6, ipadx = 0, sticky = "ns")
	for i in range(len(elements)):
		text.insert(tk.INSERT, str(str(i)+'	'+elements[i]+'\n'))

	#labeling = tk.Label(root, text="No.	PICK UP AREAS", font=("Raleway", 10))
	#labeling.grid(columnspan=3, column=1, row=2, ipadx = 100,  sticky='nw')
	#--------------------------------------

	button1=Button(root, text='BFS Search',font=('Raleway',20), command=tab2, bg="#20bebe", fg="white")
	button1.grid(column=1,row=3, sticky=W)
	button_dfs=Button(root, text='DFS Search',font=('Raleway',20), command=tab3, bg="#20bebe", fg="white")
	button_dfs.grid(column=1,row=3, sticky=E)

canvas = tk.Canvas(root,  width=600, height=300)
canvas.grid(columnspan=3, rowspan=3)
tab1()
canvas = tk.Canvas(root,  width=600, height=100)
canvas.grid(columnspan=3, rowspan=3)
root.mainloop()


